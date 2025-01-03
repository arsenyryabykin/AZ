import openpyxl
from os import path
import sys
from tkinter.messagebox import showerror, showwarning


# Проверка существования файла #
file_exist = path.exists("Стеллажи.xlsx")
if(file_exist):
    workbook = openpyxl.load_workbook("Стеллажи.xlsx")
else:
    showerror("Ошибка", "Отсутствует файл Стеллажи.xlsx")
    sys.exit()



xlsx_coords = {"А1":(3,2), "А2":(3,3), "А3":(3,4),"А4":(3,5),"А5":(3,6),"А6":(3,7),"А7":(3,8),"А8":(3,9),"А9":(3,10),"А10":(3,11),
               "Б1":(5,2),"Б2":(5,3),"Б3":(5,4),"Б4":(5,5),"Б5":(5,6),"Б6":(5,7),"Б7":(5,8),"Б8":(5,9),"Б9":(5,10),
                "В1":(7,2),"В2":(7,3),"В3":(7,4),"В4":(7,5),"В5":(7,6),"В6":(7,7),"В7":(7,8),"В8":(7,9),"В9":(7,10),"В10":(7,11),
                "Г1":(9,2),"Г2":(9,3), "Г3":(9,4),"Г4":(9,5),"Г5":(9,6),"Г6":(9,7),"Г7":(9,8),"Г8":(9,9),"Г9":(9,10),
                "Д1":(11,2),"Д2":(11,3),"Д3":(11,4),"Д4":(11,5),"Д5":(11,6),"Д6":(11,7),"Д7":(11,8),"Д8":(11,9),"Д9":(11,10),"Д10":(11,11),
                "Е1":(13,2),"Е2":(13,3),"Е3":(13,4),"Е4":(13,5),"Е5":(13,6),"Е6":(13,7),"Е7":(13,8),"Е8":(13,9),"Е9":(13,10),
                "Ж1":(15,2),"Ж2":(15,3),"Ж3":(15,4),"Ж4":(15,5),"Ж5":(15,6),"Ж6":(15,7),"Ж7":(15,8),"Ж8":(15,9),"Ж9":(15,10),"Ж10":(15,11)}

tvs_list = []
repeat_tvs_list = []
def get_shelf(shelf_number):
    shelf_i = workbook[str(shelf_number)]
    shelf = dict()

    for key, value in xlsx_coords.items():

            tvs = shelf_i.cell(row=value[0], column=value[1]).value
            if(tvs == None):
                tvs = suz = ""
            else:
                if(tvs not in tvs_list):
                    tvs_list.append(tvs)
                elif(tvs not in repeat_tvs_list):
                    repeat_tvs_list.append(tvs)  # Проверка повторов #


                suz = shelf_i.cell(row=value[0]+1, column=value[1]).value
                if(suz == None):
                    suz = ""
            shelf[key] = (tvs, suz)

    return shelf

def check_unique(shelves):
    repeatings = []
    for tvs in repeat_tvs_list:
        repeats = []
        for i, shelf in enumerate(shelves, 1): # Выбор стеллажа
            for cell in shelf:  # Проход по ячейкам i-го стеллажа
                if(cell.tvs_text == tvs):
                    repeats.append(str(cell.id) + "-" + str(i))
        repeatings.append(repeats)

    if(repeatings):
        error_text = "Совпадения ТВС в следующих ячейках:\n"
        for element in repeatings:
            error_text += " ".join(element)
            error_text += '\n'

        showerror("Ошибка", error_text)
        sys.exit()

